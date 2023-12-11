import os
import re
import pandas as pd
import openpyxl as op
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import Color


class OrderProcessor:
    def __init__(self, file_path):
        self.df = pd.read_excel(file_path, skiprows=1)
        self.result = pd.DataFrame()
        self._preprocess_data()

    #데이터 전처리 메서드
    def _preprocess_data(self):
        columns_to_keep = ['상품명', '옵션정보', '수량']
        self.df = self.df[columns_to_keep]

        self.df.drop(self.df[~self.df['상품명'].str.contains('\[치카노트\]')].index, inplace=True)
        self.df['상품명'] = self.df['상품명'].str.replace(r'\[치카노트\]\s*', '', regex=True)

        index_name = set(self.df['상품명'])
        index_name = list(index_name)

        self.df['상품명'] = self.df['상품명'].str.replace(r'\(\d+장\) / A5 \d+공 속지', '', regex=True)
        self.df['상품명'] = self.df['상품명'].str.replace(r'(\b5세트\b|\b10매\b).*$', '', regex=True)

        index_name = set(self.df['상품명'])
        index_name = list(index_name)

        self.df['옵션정보'] = self.df['옵션정보'].str.replace(r'컬러: ', '')
        self.df['옵션정보'] = self.df['옵션정보'].str.replace(r'시작', '')

        self.df['상품명'] = self.df.apply(
            lambda row: f"{row['상품명']}/ {row['옵션정보'].split(' / 종류: ')[-1]}" if '종류' in row['옵션정보'] else row['상품명'], axis=1)

        self.df['옵션정보'] = self.df['옵션정보'].str.replace(r' / 종류: .*$', '', regex=True)
        self.df['옵션정보'] = self.df['옵션정보'].str.replace(r' /.*$', '')

        self.df['상품명'] = self.df['상품명'].str.strip()

    #데이터 가공 메서드
    def process_orders(self):

        index_name = set(self.df['상품명'])
        index_name = list(index_name)
        index_name = sorted(set(self.df['상품명']))
        index_name.append("색깔별 total")

        data = {str(i): [0] * len(index_name) for i in range(1, 12)}
        data['총합'] = [0] * len(index_name)

        self.result = pd.DataFrame(data, index=index_name)
        self.result.columns = ['1.화이트', '2.크림', '3.딸기우유', '4.하이틴핑크', '5.복숭아', '6.살구', '7.레몬', '8.배추', '9.메론', '10.소다',
                               '11.타로밀크티', "총합"]

        for idx, row in self.df.iterrows():
            index = row.tolist()[0]
            column = row.tolist()[1]
            self.result.at[index, column] += int(row.tolist()[2])

        self.result['총합'] = self.result.iloc[:, :-1].sum(axis=1)
        self.result.loc['색깔별 total'] = self.result.sum()


    #엑셀 파일로 변환하는 메서드
    def save_result_to_excel(self, output_file_path='발주 수량.xlsx'):
        base_name, ext = os.path.splitext(output_file_path)
        counter = 1

        while os.path.exists(output_file_path):
            output_file_path = f"{base_name} ({counter}){ext}"
            counter += 1

        with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
            self.result.to_excel(writer, index=True, sheet_name='Sheet1')

            #엑셀 파일 꾸미는 코드들
            sheet = writer.sheets['Sheet1']

            for column_cells in sheet.columns:
                new_column_length = max(len(str(cell.value)) for cell in column_cells)
                new_column_letter = (get_column_letter(column_cells[0].column))
                if new_column_length > 0:
                    sheet.column_dimensions[new_column_letter].width = 15

            sheet.column_dimensions['A'].width = 30

            #색상 변경
            sheet['C1'].fill = PatternFill(fill_type='solid', fgColor=Color('FFF1C1'))
            sheet['D1'].fill = PatternFill(fill_type='solid', fgColor=Color('FCE6E4'))
            sheet['E1'].fill = PatternFill(fill_type='solid', fgColor=Color('FEB0CD'))
            sheet['F1'].fill = PatternFill(fill_type='solid', fgColor=Color('F0CECE'))
            sheet['G1'].fill = PatternFill(fill_type='solid', fgColor=Color('F5D3AB'))
            sheet['H1'].fill = PatternFill(fill_type='solid', fgColor=Color('FFFC78'))
            sheet['I1'].fill = PatternFill(fill_type='solid', fgColor=Color('D2DB6C'))
            sheet['J1'].fill = PatternFill(fill_type='solid', fgColor=Color('E0EAD5'))
            sheet['K1'].fill = PatternFill(fill_type='solid', fgColor=Color('CAE3EB'))
            sheet['L1'].fill = PatternFill(fill_type='solid', fgColor=Color('D5D5E1'))









